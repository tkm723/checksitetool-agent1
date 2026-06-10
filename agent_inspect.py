#!/usr/bin/env python3
# coding: utf-8
"""医療・クリニック系Webサイト検品ツール（マルチエージェント版）"""

import argparse
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urljoin, urlparse

import anthropic
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font

DEFAULT_MODEL = "claude-sonnet-4-6"

# ── Agent system prompts ──────────────────────────────────────────────────────

PROMPT_BASIC = """\
あなたはWebサイトの基本情報照合エージェントです。
医療・クリニック系Webサイトの基本情報が既存サイトと一致しているかチェックします。

## 入力データ
- page: チェック対象ページの抽出情報（JSON）
- master: 既存サイトのトップページから抽出した基本情報（JSON）

## チェック項目（トップページ is_top=true のみ）
- 電話番号：既存サイトのTELと構築サイトのTEL/TELリンクが一致するか
- FAX番号：既存サイトのFAXと構築サイトのFAXが一致するか
- 院名：既存サイトと構築サイトで院名が一致するか（一方が他方を含む場合は一致とみなす）
- 院長名：既存サイトと構築サイトで院長・ドクター名が一致するか（スペース除去・部分一致）
- 診療時間：既存サイトと構築サイトで診療時間が一致するか（表記揺れを考慮）
- 休診日：既存サイトと構築サイトで休診日が一致するか
- コピーライト：表記が一致するか・年号が古くないか（現在年と比較）
- 最寄り駅：駅名・所要時間が一致するか
- グロナビとフッターで同じhrefなのにテキストが異なるものがないか

## チェック項目（全ページ）
- TELリンク（tel:）と表示されている電話番号が一致するか
- 電話番号のハイフン統一：同じページ内でハイフンあり・なしが混在していないか

## 判定ルール
- 電話番号の比較はハイフン・スペースを除去してから行う
- 院名は「一方が他方を含む」場合は一致とみなす（略称・法人名の違い等）
- 誤検知を減らすことを優先する。確実に問題と判断できる場合のみ指摘する
- 「〜の可能性があります」等、断言を避けた表現を使う
- ヘッダー・フッター起因の問題はトップページのみ報告する

## 出力形式
JSON配列のみを返す。余分な説明文・マークダウンは一切不要。
[
  {
    "type": "ng",
    "category": "電話番号",
    "text": "電話番号が既存サイトと一致しない可能性があります",
    "existing": "既存サイトの値",
    "found": "構築サイトの値",
    "page_url": "https://example.com/",
    "page_path": "/"
  }
]
問題がなければ [] を返す。\
"""

PROMPT_SEO = """\
あなたはWebサイトのSEO・メタ情報チェックエージェントです。
医療・クリニック系Webサイトのメタ情報・見出し構造・SEO設定を確認します。

## 入力データ
- page: チェック対象ページの抽出情報（JSON）

## チェック項目（トップページ is_top=true のみ）
- titleタグ：空でないか
- meta description：設定されているか
- meta keywords：設定されているか
- og:title / og:description / og:image：設定されているか
- canonicalタグ：設定されているか
- noindex：意図しない設定がされていないか（staging環境のnoindexが残っていないか）
- JSON-LD構造化データ：存在するか

## チェック項目（全ページ）
- h1タグ：存在するか・2つ以上重複していないか
- h1にダミーワード（「テスト」「テキスト」「ダミー」「サンプル」等）が入っていないか
- 見出し階層：h1→h2→h3の順番が守られているか（2レベル以上スキップはNG）
- パンくずリスト（トップ以外）：末尾とh1タイトルが一致するか・存在するか

## 判定ルール
- 見出し階層は「2つ以上レベルが飛ぶ」場合のみ指摘（h1→h3はNG、h1→h2→h2はOK）
- noindexはstagingサイトに残っている場合が多いため必ず確認
- パンくずが見つからない場合は type="manual"（要手動確認）で報告

## 出力形式
JSON配列のみを返す。余分な説明文・マークダウンは一切不要。
[{"type": "ng", "category": "見出し", "text": "h1タグが2個あります", "found": "...", "page_url": "...", "page_path": "..."}]
問題がなければ [] を返す。\
"""

PROMPT_DUMMY = """\
あなたはWebサイトのダミーコンテンツ検出エージェントです。
医療・クリニック系Webサイトに残存するダミーテキスト・ダミー画像・ダミー料金を検出します。

## 入力データ
- page: チェック対象ページの抽出情報（JSON）。body_text・headings・images等が含まれる。

## ダミーテキスト検出パターン

### 確実なダミーキーワード（含まれていたら即NG）
「テキストテキスト」「テキストが入ります」「テキストが入る」「本文が入ります」
「見出しが入ります」「タイトルが入ります」「コンテンツが入ります」「内容が入ります」
「文章が入ります」「説明が入ります」「キャッチが入ります」「サンプルテキスト」
「ダミーテキスト」「Lorem ipsum」「NOW PRINTING」「NO IMAGE」「COMING SOON」
「準備中の画像」「テスト文章」「テストテキスト」

### プレースホルダーパターン
- 「〇〇〇」「●●●」「□□□」等の記号3文字以上の連続
- 「〇〇回」「〇〇円」「〇〇ヶ月」等の丸文字+単位
- 「〇〇クリニック」「〇〇医院」等の丸文字施設名
- 「山田太郎」「田中太郎」「田中花子」「John Doe」等のダミー人名
- 「〇〇大学」「○○年　〇〇」等のダミー経歴

### ダミー料金
- 全桁ゼロの金額（「00,000円」「0,000円」等）
- 「〇〇〇円」等の丸文字金額
- 料金の片方だけがゼロ（「10,000円〜00,000円」等）

### ダミー画像（imagesフィールドを確認）
- src に「pixta」「placeholder」「dummy」「sample」「noimage」「now_printing」等を含む
- altテキストに「ダミー」「サンプル」「準備中」「テスト」等を含む

## 重要な注意事項
- 同じダミーテキストが複数箇所あれば全件報告する（上限50件）
- 本文中の正当なテキストを誤検知しないよう注意（「テキスト」単体はOK、「テキストテキスト」はNG）
- header/footer/nav内のダミーは対象外

## 出力形式
JSON配列のみを返す。余分な説明文・マークダウンは一切不要。
[{
  "type": "ng",
  "category": "ダミーテキスト",
  "text": "ダミーテキストが残っています",
  "found": "「テキストテキスト」",
  "heading": "h4:「e-max」> 表の行「料金」",
  "page_url": "...",
  "page_path": "..."
}]
問題がなければ [] を返す。\
"""

PROMPT_LINK_UI = """\
あなたはWebサイトのリンク・UIチェックエージェントです。
リンクの設定ミス・画像のalt属性・フォーム設定等を確認します。

## 入力データ
- page: チェック対象ページの抽出情報（JSON）

## チェック項目（全ページ）

### 外部リンク
- external_linksフィールドを確認し、target="_blank"が設定されていないものを指摘

### 空リンク・無効リンク
- invalid_linksフィールドを確認
- ページトップボタン（#top, #pagetop等）は除外
- ヘッダー・フッター内の空リンクはトップページのみ報告

### alt属性
- imagesフィールドで alt=null（未設定）の画像を指摘
- alt=""（明示的な空設定）はOK
- alt="image" "img" "photo" "写真" "画像" 等の無意味な値もNG
- /header /footer /logo favicon icon blank を含むsrcは除外

### reCAPTCHA
- has_form=true のページで has_recaptcha=false の場合はNG

## チェック項目（トップページ以外 is_top=false）
- パンくずリスト：breadcrumbフィールドがnullまたは空なら type="manual" で「要手動確認」
- パンくず末尾とh1が不一致なら type="ng" で指摘

## 出力形式
JSON配列のみを返す。余分な説明文・マークダウンは一切不要。
[{"type": "ng"/"manual", "category": "外部リンク"/"空リンク"/"alt未設定"/"reCAPTCHA"/"パンくず",
  "text": "...", "found": "...", "page_url": "...", "page_path": "..."}]
問題がなければ [] を返す。\
"""

PROMPT_MEDICAL_GL = """\
あなたは医療広告ガイドライン専門のチェックエージェントです。
日本の医療広告ガイドラインに基づき、医療・クリニック系Webサイトのコンテンツをチェックします。

## 入力データ
- page: チェック対象ページの抽出情報（JSON）。body_text・headings・path等が含まれる。

## チェック対象外ページ（以下のpathは医療広告チェックをスキップ）
/recruit /career /privacy /policy /access

## チェック項目（全ページ）

### 最上級・断定表現（違反）
以下の表現が含まれる場合は指摘：
「日本一」「No.1」「NO.1」「最高」「最先端」「最新鋭」「唯一」「他院では」
「絶対安全」「100%」「必ず」「確実に」「副作用なし」「痛みゼロ」
「〜を保証」「効果を約束」「治ります」「完治します」

### 体験談・口コミ（原則禁止）
患者の感想・体験談・口コミに見える表現（「〜でよかった」「先生のおかげで」等）

### 料金記載があるページ（自由診療の必要記載事項）
料金（数字+円）の記載がある場合、以下が揃っているか確認：
- 治療内容の説明
- 治療期間または回数
- リスク・副作用の記載
- 税込表示

料金表示：「税抜」「税別」「+税」等の税抜き表示はNG

### URLに /case /jirei /before /after /works を含む、または見出しに「症例」「ビフォーアフター」を含む
症例ページの必要記載（4項目全てが揃っているか）：
① 治療内容
② 治療期間または回数
③ 費用
④ リスク・副作用

### 未承認医薬品ページ（「未承認」「個人輸入」のキーワードがある場合）
5項目全ての記載確認：
① 未承認医薬品・医療機器である旨
② 入手経路
③ 国内の同一成分を含む承認医薬品等の有無
④ 諸外国における安全性情報
⑤ 医薬品副作用被害救済制度の対象外である旨

### テキスト品質（AI判定）
- 明らかな誤字・脱字
- 表記ゆれ（「ホワイトニング」と「ホワイニング」等）

## 判定ルール
- 確実に違反と判断できるもののみ指摘する（グレーゾーンは type="manual"）
- 「〜の可能性があります」等、断定を避けた表現を使う

## 出力形式
JSON配列のみを返す。余分な説明文・マークダウンは一切不要。
[{"type": "ng"/"manual", "category": "医療広告GL"/"自由診療"/"症例ページ"/"未承認医薬品"/"テキスト品質",
  "text": "...", "found": "...", "page_url": "...", "page_path": "..."}]
問題がなければ [] を返す。\
"""

# ── URL helpers ───────────────────────────────────────────────────────────────

MEDIA_RE = re.compile(
    r'\.(jpg|jpeg|png|gif|webp|svg|pdf|zip|mp4|mp3|doc|docx|xls|xlsx)(\?.*)?$', re.I
)
PHONE_RE = re.compile(r'0\d{1,4}[-－\-]?\d{1,4}[-－\-]?\d{4}')


def _staging_base(url: str) -> str:
    parts = urlparse(url).path.strip('/').split('/')
    if len(parts) >= 2 and '.' in parts[1]:
        return '/' + '/'.join(parts[:2])
    return ''


def _rel_path(url: str, base_path: str) -> str:
    path = urlparse(url).path
    if base_path and path.startswith(base_path):
        path = path[len(base_path):]
    return path or '/'


def _is_skip(url: str, base_path: str) -> bool:
    path = _rel_path(url, base_path)
    if MEDIA_RE.search(url):
        return True
    if re.search(r'/wp-admin/|/page/\d+/?$|\d{4}/\d{2}/\d{2}/', url):
        return True
    if re.search(r'/(blog|news|case)/?$', path):
        return True
    if re.search(r'/(blog|news)/[^/]+/?$', path):
        return True
    if 'sitemap' in url.lower():
        return True
    return False

# ── HTML extractor (Agent 1 — Python) ────────────────────────────────────────

def extract_page_info(url: str, soup: BeautifulSoup, base_path: str) -> dict:
    path = _rel_path(url, base_path)
    own = urlparse(url).netloc

    # Body text (strip scripts/styles first)
    for tag in soup(['script', 'style']):
        tag.decompose()
    body_text = soup.get_text(separator=' ', strip=True)[:5000]

    # Headings
    headings = [
        {'level': int(h.name[1]), 'text': h.get_text(strip=True)[:100]}
        for h in soup.find_all(['h1','h2','h3','h4','h5','h6'])
    ][:50]

    # Meta
    title_tag = soup.find('title')
    title = title_tag.get_text(strip=True) if title_tag else ''

    def meta_content(name_val):
        t = soup.find('meta', attrs={'name': re.compile(f'^{name_val}$', re.I)})
        return (t.get('content') or '').strip() if t else ''

    def og_content(prop):
        t = soup.find('meta', property=re.compile(f'^og:{prop}$', re.I))
        return (t.get('content') or '').strip() if t else ''

    canonical_tag = soup.find('link', rel='canonical')
    canonical = canonical_tag.get('href', '') if canonical_tag else ''

    robots = soup.find('meta', attrs={'name': re.compile(r'^robots$', re.I)})
    noindex = bool(robots and 'noindex' in (robots.get('content') or '').lower())

    # Phones & TEL links
    phones = list(dict.fromkeys(PHONE_RE.findall(body_text)))[:10]
    tel_links = [
        {'href': a['href'], 'text': a.get_text(strip=True)}
        for a in soup.find_all('a', href=re.compile(r'^tel:', re.I))
    ][:10]

    # FAX
    fax = re.findall(
        r'(?:FAX|Fax|fax|ファックス)[：: ]*(' + r'0\d{1,4}[-－\-]?\d{1,4}[-－\-]?\d{4})',
        body_text
    )[:5]

    # Postal codes
    postal = re.findall(r'〒?\d{3}[-\-]\d{4}', body_text)[:5]

    # Copyright (footer, excluding nav)
    copyright_text = ''
    footer = soup.find('footer')
    if footer:
        fn = footer.find('nav')
        ft = footer.get_text()
        if fn:
            ft = ft.replace(fn.get_text(), '')
        m = re.search(r'(?:©|Copyright)[^\n]*', ft)
        if m:
            copyright_text = m.group().strip()[:80]

    # Nearest station
    station = ''
    access_el = soup.find(id=re.compile(r'^access$', re.I)) or \
                soup.find(class_=re.compile(r'\baccess\b', re.I))
    if access_el:
        m = re.search(r'([^\s]{2,10}駅)[^線]{0,20}(徒歩|車|バス)[^\d]{0,5}(\d+)分', access_el.get_text())
        if m:
            station = m.group()

    # Images
    images = [
        {'src': img.get('src', '')[:100], 'alt': img.get('alt')}
        for img in soup.find_all('img')
    ][:50]

    # External links
    ext_links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        try:
            abs_href = urljoin(url, href)
            link_netloc = urlparse(abs_href).netloc
        except Exception:
            continue
        if link_netloc and link_netloc != own:
            ext_links.append({
                'href': href[:100],
                'text': a.get_text(strip=True)[:40],
                'target': a.get('target', ''),
            })
    ext_links = ext_links[:30]

    # Invalid/empty links
    invalid_links = []
    for a in soup.find_all('a', href=True):
        if a.find_parent('nav'):
            continue
        href = a['href'].strip()
        if href in ('', '#', 'javascript:void(0)', 'javascript:void(0);', 'javascript:;'):
            cls = ' '.join(a.get('class', []))
            if re.search(r'pagetop|page-top|totop|scroll', cls + href, re.I):
                continue
            invalid_links.append({'href': href, 'text': a.get_text(strip=True)[:30]})
    invalid_links = invalid_links[:20]

    # Nav and footer link maps
    def link_map(el):
        if not el:
            return {}
        return {a['href']: a.get_text(strip=True)
                for a in el.find_all('a', href=True) if a.get_text(strip=True)}

    navs = soup.find_all('nav')
    nav_map = link_map(navs[0]) if navs else {}
    footer_map = link_map(footer) if footer else {}

    # Breadcrumb
    bc_el = (
        soup.find(class_=re.compile(r'breadcrumb', re.I)) or
        soup.find(id=re.compile(r'breadcrumb', re.I)) or
        soup.find('nav', attrs={'aria-label': re.compile(r'breadcrumb|パンくず', re.I)}) or
        soup.find(class_=re.compile(r'\bpan\b|pankuzu', re.I))
    )
    breadcrumb = [el.get_text(strip=True) for el in bc_el.find_all(['li','span','a'])] \
        if bc_el else None

    # Form / reCAPTCHA
    has_form = bool(soup.find('form'))
    html_lower = str(soup).lower()
    has_recaptcha = 'recaptcha' in html_lower or 'grecaptcha' in html_lower

    # JSON-LD
    has_jsonld = bool(soup.find('script', type='application/ld+json'))

    return {
        'url': url,
        'path': path,
        'is_top': path in ('/', ''),
        'title': title,
        'meta_desc': meta_content('description'),
        'meta_keywords': meta_content('keywords'),
        'ogp': {
            'title': og_content('title'),
            'description': og_content('description'),
            'image': og_content('image'),
        },
        'canonical': canonical,
        'noindex': noindex,
        'has_jsonld': has_jsonld,
        'headings': headings,
        'phones': phones,
        'tel_links': tel_links,
        'fax': fax,
        'postal': postal,
        'copyright': copyright_text,
        'station': station,
        'external_links': ext_links,
        'invalid_links': invalid_links,
        'images': images,
        'nav_map': dict(list(nav_map.items())[:30]),
        'footer_map': dict(list(footer_map.items())[:30]),
        'breadcrumb': breadcrumb,
        'has_form': has_form,
        'has_recaptcha': has_recaptcha,
        'body_text': body_text,
    }

# ── Crawler ───────────────────────────────────────────────────────────────────

def crawl(target_url: str, base_path: str) -> dict:
    netloc = urlparse(target_url).netloc
    visited: set = set()
    queue: list = [target_url]
    pages: dict = {}
    session = requests.Session()
    session.headers['User-Agent'] = 'Mozilla/5.0 (compatible; WebInspector/1.0)'

    while queue:
        url = queue.pop(0)
        if url in visited:
            continue
        if urlparse(url).netloc != netloc:
            continue
        if _is_skip(url, base_path):
            visited.add(url)
            continue

        visited.add(url)
        try:
            resp = session.get(url, timeout=15, allow_redirects=True)
            if 'html' not in resp.headers.get('content-type', ''):
                continue
            soup = BeautifulSoup(resp.text, 'lxml')
            pages[url] = {'soup': soup, 'status': resp.status_code}

            for a in soup.find_all('a', href=True):
                href = a['href'].strip()
                if not href or href.startswith(('mailto:', 'tel:', 'javascript:')):
                    continue
                abs_url = urljoin(url, href).split('#')[0].split('?')[0]
                if abs_url not in visited and urlparse(abs_url).netloc == netloc:
                    queue.append(abs_url)

        except Exception as e:
            print(f'  エラー: {url} — {e}', file=sys.stderr)
            pages[url] = {'soup': None, 'status': 0}

    return pages

# ── LLM agent helpers ─────────────────────────────────────────────────────────

def _parse_json_response(text: str) -> list:
    text = re.sub(r'^```(?:json)?\s*', '', text.strip(), flags=re.M)
    text = re.sub(r'```\s*$', '', text.strip(), flags=re.M)
    text = text.strip()
    m = re.search(r'\[[\s\S]*\]', text)
    if m:
        return json.loads(m.group())
    return json.loads(text)


def run_agent(client: anthropic.Anthropic, name: str, system_prompt: str,
              payload: dict, model: str) -> list:
    try:
        resp = client.messages.create(
            model=model,
            max_tokens=4096,
            system=system_prompt,
            messages=[{'role': 'user', 'content': json.dumps(payload, ensure_ascii=False)}],
        )
        return _parse_json_response(resp.content[0].text)
    except Exception as e:
        page_url  = (payload.get('page') or {}).get('url', '')
        page_path = (payload.get('page') or {}).get('path', '')
        return [{'type': 'error', 'category': f'エージェントエラー({name})',
                 'text': str(e), 'found': '', 'page_url': page_url, 'page_path': page_path}]


def check_page(client: anthropic.Anthropic, page_info: dict,
               master_info: dict, model: str) -> list:
    agents = [
        ('基本情報チェッカー',   PROMPT_BASIC),
        ('SEOチェッカー',       PROMPT_SEO),
        ('ダミーチェッカー',     PROMPT_DUMMY),
        ('リンク・UIチェッカー', PROMPT_LINK_UI),
        ('医療広告GLチェッカー', PROMPT_MEDICAL_GL),
    ]
    payload = {'page': page_info, 'master': master_info}

    all_issues: list = []
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {
            ex.submit(run_agent, client, name, prompt, payload, model): name
            for name, prompt in agents
        }
        for future in as_completed(futures):
            name = futures[future]
            try:
                issues = future.result()
                if isinstance(issues, list):
                    all_issues.extend(issues)
            except Exception as e:
                print(f'  {name} 例外: {e}', file=sys.stderr)

    # Dedup by (category, text[:60], page_path)
    seen: set = set()
    deduped: list = []
    for issue in all_issues:
        key = (issue.get('category', ''), issue.get('text', '')[:60], issue.get('page_path', ''))
        if key not in seen:
            seen.add(key)
            deduped.append(issue)
    return deduped

# ── Manual check items (top page only) ───────────────────────────────────────

MANUAL_ITEMS = [
    '【レイアウト】デザインカンプとの照合・レイアウト崩れ（PC/SP/iPad）・WEBフォント確認',
    '【ダミー画像】グレーの四角のみのダミー画像（テキストなし）は自動検知不可のため全ページ目視確認',
    '【投稿】テスト投稿が削除されているか・投稿一覧・アイキャッチの表示崩れ',
    '【フォーム】テスト送信・必須項目・自動返信メール・reCAPTCHA・完了ページへの遷移',
    '【その他】Googleマップのピン位置・404ページ・ページトップボタン・JSの動作・画像WebP変換',
    '【基本情報】設備・器具等の名称・最寄り駅からの所要時間がサイト内で統一されているか',
]

# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(page_results: dict, existing_url: str, out_path: str):
    wb = openpyxl.Workbook()
    FILL_NG   = PatternFill(fill_type='solid', fgColor='FFCCCC')
    FILL_WARN = PatternFill(fill_type='solid', fgColor='FFFACC')
    FILL_OK   = PatternFill(fill_type='solid', fgColor='CCFFCC')
    FILL_HDR  = PatternFill(fill_type='solid', fgColor='4472C4')
    FONT_HDR  = Font(color='FFFFFF', bold=True)

    def fill_for(t):
        return FILL_NG if t == 'ng' else FILL_WARN if t in ('manual', 'error') else FILL_OK

    def kind_label(t):
        return {'ng': 'NG', 'manual': '要手動確認', 'error': 'エラー'}.get(t, 'OK')

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = 'サマリー'
    for c, h in enumerate(['ページパス', 'NG件数', '要手動確認件数', '主なNGカテゴリ'], 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR

    for r, (url, issues) in enumerate(page_results.items(), 2):
        path  = _rel_path(url, '')
        ng_n  = sum(1 for x in issues if x.get('type') == 'ng')
        man_n = sum(1 for x in issues if x.get('type') == 'manual')
        cats  = list(dict.fromkeys(x.get('category','') for x in issues if x.get('type') == 'ng'))[:3]
        fill  = FILL_NG if ng_n else FILL_WARN if man_n else FILL_OK
        for c, v in enumerate([path, ng_n, man_n, ', '.join(cats)], 1):
            ws1.cell(row=r, column=c, value=v).fill = fill

    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 45

    # Sheet 2: Detail
    ws2 = wb.create_sheet('ページ別詳細')
    headers = ['ページパス', '構築URL', '既存URL', '種別', 'カテゴリ', '内容', '既存サイトの値', '検出値', '箇所']
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR

    row = 2
    base_ex = existing_url.rstrip('/')
    for url, issues in page_results.items():
        path = _rel_path(url, '')
        ex_url = base_ex + (path if path != '/' else '/')
        for x in issues:
            vals = [
                path, url, ex_url,
                kind_label(x.get('type', '')),
                x.get('category', ''),
                x.get('text', ''),
                x.get('existing', ''),
                x.get('found', ''),
                x.get('heading', x.get('location', '')),
            ]
            f = fill_for(x.get('type', ''))
            for c, v in enumerate(vals, 1):
                ws2.cell(row=row, column=c, value=v).fill = f
            row += 1

    import openpyxl.utils as xu
    for c, w in enumerate([35, 55, 55, 12, 20, 70, 30, 40, 35], 1):
        ws2.column_dimensions[xu.get_column_letter(c)].width = w

    wb.save(out_path)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='医療・クリニック系Webサイト検品ツール（マルチエージェント版）')
    parser.add_argument('--existing', required=True, help='既存サイトURL（正解データ）')
    parser.add_argument('--target',   required=True, help='構築サイトURL（チェック対象）')
    parser.add_argument('--model',    default=DEFAULT_MODEL,
                        help=f'使用Claudeモデル (default: {DEFAULT_MODEL})')
    args = parser.parse_args()

    target_url   = args.target.rstrip('/')
    existing_url = args.existing.rstrip('/')
    base_path    = _staging_base(target_url)

    if base_path:
        print(f'ステージングURL検出: ベースパス = {base_path}')

    client = anthropic.Anthropic()  # ANTHROPIC_API_KEY 環境変数から読み込み

    # 1. Crawl target site
    print(f'\n構築サイトの巡回を開始: {target_url}')
    pages = crawl(target_url, base_path)
    print(f'巡回完了: {len(pages)} ページ\n')

    # 2. Fetch existing site top page for master_info
    master_info: dict = {}
    print(f'既存サイトのトップページを取得: {existing_url}')
    try:
        sess = requests.Session()
        sess.headers['User-Agent'] = 'Mozilla/5.0 (compatible; WebInspector/1.0)'
        resp = sess.get(existing_url, timeout=15)
        e_soup = BeautifulSoup(resp.text, 'lxml')
        master_info = extract_page_info(existing_url, e_soup, '')
        print('取得完了\n')
    except Exception as e:
        print(f'既存サイト取得エラー: {e}\n', file=sys.stderr)

    # 3. Check all pages
    page_results: dict = {}
    for url, data in pages.items():
        if not data.get('soup'):
            path = _rel_path(url, base_path)
            page_results[url] = [{
                'type': 'error', 'category': 'クロール',
                'text': f'ページ取得失敗 (status={data.get("status","?")})',
                'found': '', 'page_url': url, 'page_path': path,
            }]
            continue

        page_info = extract_page_info(url, data['soup'], base_path)
        path = page_info['path']
        print(f'エージェント起動中: {path} …', end=' ', flush=True)

        issues = check_page(client, page_info, master_info, args.model)

        # Append manual items on top page
        if page_info['is_top']:
            for text in MANUAL_ITEMS:
                issues.append({
                    'type': 'manual', 'category': '手動確認', 'text': text,
                    'found': '', 'page_url': url, 'page_path': path,
                })

        page_results[url] = issues
        ng_n  = sum(1 for x in issues if x.get('type') == 'ng')
        man_n = sum(1 for x in issues if x.get('type') == 'manual')
        print(f'NG:{ng_n}件  要確認:{man_n}件')

    # 4. Write Excel
    ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = f'検品結果_{ts}.xlsx'
    print(f'\n結果を出力中: {out}')
    write_excel(page_results, existing_url, out)

    total_ng  = sum(sum(1 for x in v if x.get('type') == 'ng')     for v in page_results.values())
    total_man = sum(sum(1 for x in v if x.get('type') == 'manual') for v in page_results.values())
    print(f'\n完了  NG={total_ng}件  要手動確認={total_man}件')
    print(f'出力ファイル: {out}')


if __name__ == '__main__':
    main()
